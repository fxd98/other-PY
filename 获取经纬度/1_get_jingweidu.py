from osgeo import gdal
from pylab import *  # 支持中文
mpl.rcParams['font.sans-serif'] = ['SimHei']
from openpyxl import Workbook
import os,numpy as np,pandas as pd

# 创建一个Workbook对象
work = Workbook()

def get_data(path):
    print('正在处理：',path)
    dataset = gdal.Open(path)  # 打开tif

    geo_information = dataset.GetGeoTransform()
    A1 = path
    B1 = dataset.RasterYSize
    C1 = dataset.RasterXSize
    D1 = geo_information[0]
    E1 = geo_information[3]
    F1 = geo_information[1]
    G1 = geo_information[5]
    return [A1,B1,C1,D1,E1,F1,G1]

def out(data, name):
    ws = work.active
    ws['A1'] = '位置'
    ws['B1'] = '图片高度'
    ws['C1'] = '图片宽度'
    ws['D1'] = '左上角经度'
    ws['E1'] = '左上角维度'
    ws['F1'] = 'w - e像素分辨率 / 像素宽度'
    ws['G1'] = 'n - s像素分辨率 / 像素高度（北半球上图像为负值）'
    for i in range(len(data)):
        rows = []
        row_length = len(data[i])
        if row_length != 0:
            for j in range(row_length):
                # rows.append(data[i][j])
                rows.append(data[i])
                ws.append([rows[j]])
        print(rows)
    work.save(name)

path = 'D:/FuXiaodi/####/##/BIJIE_ALL/NEW_FENGE/TIF/'
files = os.listdir(path)

all_data = []
for file in files:
    if file[-4:] == '.tif':
        all_data.append(get_data(path + file))
        
    else:
        new_path1 = path + file
        for file1 in os.listdir(new_path1):
            if file1[-4:] == '.tif':
                all_data.append(get_data(new_path1 + '/'+ file1))
                
            else:
                new_path2 = new_path1 + '/' + file1
                for file2 in os.listdir(new_path2):
                    if file2[-4:] == '.tif' :
                        all_data.append(get_data(new_path2 + '/' + file2))
                        
                    else:
                        new_path3 = new_path2 + '/' + file2
                        for file3 in os.listdir(new_path3):
                            if file3[-4:] == '.tif' :
                                all_data.append(get_data(new_path3 + '/' + file3))
                                
# out(all_data, '所有图片信息.xlsx')
pa = pd.DataFrame(all_data,columns=['位置','图片高度','图片宽度','左上角经度','左上角维度',
                                    'w - e像素分辨率 / 像素宽度','n - s像素分辨率 / 像素高度（北半球上图像为负值）'])
# pa.to_excel("毕节市--图片信息.xlsx")
# print('表已经生成')